# app.py
# Tokenomics Audit AI — all-in-one deployable Streamlit app
# Includes: TDeFi styling, sheet template download/upload, metrics, grounded AI, PDF export

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from fpdf import FPDF
import tempfile
import os
import re
import io
import json
import base64
from openai import OpenAI

from typing import Optional

# -----------------------------
# Helper: Safe OpenAI API key retrieval
# -----------------------------
def get_openai_api_key():
    env_key = os.environ.get("OPENAI_API_KEY")
    if env_key:
        return env_key
    try:
        if hasattr(st, "secrets") and hasattr(st.secrets, "__getitem__"):
            openai_section = st.secrets.get("openai") if hasattr(st.secrets, "get") else None
            if openai_section and isinstance(openai_section, dict):
                return openai_section.get("api_key")
    except Exception:
        pass
    return None

# -----------------------------
# Streamlit page config + CSS
# -----------------------------
st.set_page_config(
    page_title="Tokenomics Audit AI",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    .company-logo {
        position: relative;
        background: transparent;
        padding: 0;
        border: none;
        box-shadow: none;
        max-width: 220px;
        max-height: 80px;
        margin: 10px 0 0 0;
        display: block;
    }
    .company-logo img { max-width: 100%; max-height: 36px; height: auto; object-fit: contain; display: block; }

    .main-header {
        background: linear-gradient(90deg, #ffc107 0%, #fd7e14 100%);
        padding: 0.6rem 1rem;
        border-radius: 10px;
        margin-bottom: 1.2rem;
        color: white;
        box-shadow: 0 4px 15px rgba(255, 193, 7, 0.3);
        display: block;
    }
    .main-header-inner { display: flex; align-items: center; justify-content: flex-end; gap: 0.6rem; width: 100%; }
    .main-title { display: none; }
    .powered-wrap { display: none; }

    .tdefi-powered {
        text-align: center;
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 100%);
        color: white;
        font-size: 1.1rem;
        font-weight: bold;
        margin: 2rem 0 1rem 0;
        padding: 1.2rem;
        border-radius: 12px;
        border: 3px solid #ffc107;
        box-shadow: 0 8px 25px rgba(255, 193, 7, 0.2);
        position: relative;
        overflow: hidden;
    }
    .tdefi-powered::before {
        content: '';
        position: absolute;
        top: 0; left: -100%;
        width: 100%; height: 100%;
        background: linear-gradient(90deg, transparent, rgba(255, 193, 7, 0.1), transparent);
        animation: shine 3s infinite;
    }
    @keyframes shine { 0% { left: -100%; } 100% { left: 100%; } }

    .tdefi-logo-text {
        display: inline-block;
        background: linear-gradient(90deg, #ffc107 0%, #fd7e14 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-size: 1.4rem;
        font-weight: 900;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }

    .stButton > button {
        background: linear-gradient(90deg, #FF6B35 0%, #FFD700 100%);
        color: white;
        border: none;
        border-radius: 25px;
        padding: 0.5rem 2rem;
        font-weight: bold;
        box-shadow: 0 4px 15px rgba(255, 107, 53, 0.3);
        transition: all 0.3s ease;
    }
    .stButton > button:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(255, 107, 53, 0.4); }

    .metric-card {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 100%);
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #FF6B35;
        color: white;
        margin: 0.5rem 0;
    }
    .success-box {
        background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
        color: white;
        padding: 0.6rem 0.8rem;
        border-radius: 10px;
        text-align: center;
        font-weight: 700;
        font-size: 0.95rem;
    }
    .warning-box {
        background: linear-gradient(135deg, #ffc107 0%, #fd7e14 100%);
        color: white;
        padding: 1rem;
        border-radius: 10px;
        text-align: center;
        font-weight: bold;
    }

    /* Analysis sections styling for consistent alignment */
    .analysis-section { background: #111; border: 1px solid #2a2a2a; border-radius: 10px; padding: 1rem; margin: 1rem 0; }
    .analysis-title { margin: 0 0 0.25rem 0; font-weight: 800; font-size: 1.1rem; color: #ffd166; }
    .analysis-subtitle { margin: 0 0 0.5rem 0; color: #333333; font-style: italic; }
    .analysis-stat { margin: 0.25rem 0; font-weight: 700; color: #000000; }
    .analysis-price { margin: 0.35rem 0; color: #000000; }
    .analysis-price b { color: #ffb703; }
    .analysis-body { margin: 0.25rem 0; color: #000000; }
    .analysis-suggestions-title { margin: 0.5rem 0 0.25rem 0; font-weight: 700; color: #ffffff; }
    .analysis-suggestions { margin: 0.25rem 0 0.5rem 1rem; }
    .analysis-suggestions li { margin: 0.15rem 0; }
</style>
""", unsafe_allow_html=True)

# -----------------------------
# Logo helper
# -----------------------------
def _get_logo_data_uri():
    candidates = [
        "your_logo.png", "your_logo.jpeg", "your_logo.jpg",
        "logo.png", "logo.jpg", "logo.jpeg",
    ]
    for filename in candidates:
        try:
            if os.path.exists(filename):
                mime = "image/png" if filename.lower().endswith(".png") else "image/jpeg"
                with open(filename, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode("ascii")
                return f"data:{mime};base64,{b64}"
        except Exception:
            continue
    return None

_logo_uri = _get_logo_data_uri()
st.markdown(
    f"""
<div class="main-header">
  <div class="main-header-inner">
    {f'<img src="{_logo_uri}" alt="Company Logo" style="height: 31px;" />' if _logo_uri else ''}
  </div>
</div>
""",
    unsafe_allow_html=True,
)

# -----------------------------
# Template Builder (XLSX with sample rows) + CSV
# -----------------------------
def build_template_excel() -> Optional[bytes]:
    """
    Build an XLSX template if an Excel writer engine is available.
    Falls back to returning None when neither 'openpyxl' nor 'xlsxwriter' is installed.
    """
    # Prepare dataframes
    tokenomics_cols = [
        "Pool Name", "Allocation %", "Category", "TGE Unlock %",
        "Cliff (months)", "Vesting (months)", "Sellable at TGE (Yes/No)", "Notes",
    ]
    sample_rows = [
        ["Private 1", 8, "VC", 5, 6, 12, "No", "5% at TGE, 6M cliff, 12M vest"],
        ["Private 2", 6, "VC", 10, 3, 9, "No", "10% TGE, 3M cliff, 9M vest"],
        ["Public Sale", 6, "Community", 20, 0, 6, "Yes", "20% TGE, rest linear 6M"],
        ["Team & Advisors", 16, "Team", 0, 12, 18, "No", "Cliff 12M, vest 18M"],
        ["Treasury", 10, "Other", 20, 0, 48, "No", "20% at TGE, vest 48M"],
        ["Foundation", 5, "Other", 10, 0, 12, "No", "10% TGE, vest 12M"],
        ["Liquidity", 10, "Liquidity", 30, 0, 24, "Yes", "30% at TGE, vest 24M"],
        ["Community Airdrop", 4, "Community", 50, 3, 1, "Yes", "50% TGE, remaining at month 4"],
        ["Community Engagement", 30, "Community", 0, 3, 60, "No", "3M cliff, vest 60M"],
        ["Community Marketing", 5, "Community", 10, 3, 60, "Yes", "10% TGE, 3M cliff, vest 60M"],
    ]
    df_inputs = pd.DataFrame(sample_rows, columns=tokenomics_cols)

    project_cols = [
        "Project Name", "Total Supply", "TGE Price (USD)", "Liquidity Fund (USD)",
        "Project Category", "Current User Base", "User Growth %",
        "Monthly Incentive Spend (USD)", "Annual Revenue (USD)",
    ]
    df_proj = pd.DataFrame([[
        "MyProject", 1_000_000_000, 0.02, 500_000,
        "Gaming", 10_000, 10.0, 50_000, 1_000_000,
    ]], columns=project_cols)

    # Detect an available engine
    engine = None
    try:
        import openpyxl  # noqa: F401
        engine = "openpyxl"
    except Exception:
        try:
            import xlsxwriter  # noqa: F401
            engine = "xlsxwriter"
        except Exception:
            engine = None

    if engine is None:
        return None

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine=engine) as writer:
        df_inputs.to_excel(writer, index=False, sheet_name="Tokenomics Inputs")
        df_proj.to_excel(writer, index=False, sheet_name="Project Info")
    buf.seek(0)
    return buf.read()

# Helper to build CSV template DataFrame (always works)
def build_template_csv_df() -> pd.DataFrame:
    tokenomics_cols = [
        "Pool Name", "Allocation %", "Category", "TGE Unlock %",
        "Cliff (months)", "Vesting (months)", "Sellable at TGE (Yes/No)", "Notes",
    ]
    sample_rows = [
        ["Private 1", 8, "VC", 5, 6, 12, "No", "5% at TGE, 6M cliff, 12M vest"],
        ["Private 2", 6, "VC", 10, 3, 9, "No", "10% TGE, 3M cliff, 9M vest"],
        ["Public Sale", 6, "Community", 20, 0, 6, "Yes", "20% TGE, rest linear 6M"],
        ["Team & Advisors", 16, "Team", 0, 12, 18, "No", "Cliff 12M, vest 18M"],
        ["Treasury", 10, "Other", 20, 0, 48, "No", "20% at TGE, vest 48M"],
        ["Foundation", 5, "Other", 10, 0, 12, "No", "10% TGE, vest 12M"],
        ["Liquidity", 10, "Liquidity", 30, 0, 24, "Yes", "30% TGE, vest 24M"],
        ["Community Airdrop", 4, "Community", 50, 3, 1, "Yes", "50% TGE, remaining at month 4"],
        ["Community Engagement", 30, "Community", 0, 3, 60, "No", "3M cliff, vest 60M"],
        ["Community Marketing", 5, "Community", 10, 3, 60, "Yes", "10% TGE, 3M cliff, vest 60M"],
    ]
    return pd.DataFrame(sample_rows, columns=tokenomics_cols)

st.subheader("📥 Sheet-Based Inputs (optional)")
c1, c2 = st.columns([2, 3])
with c1:
    st.caption("Download a ready-made template, fill it, then upload below.")
    xl_bytes = build_template_excel()
    csv_df = build_template_csv_df()

    if xl_bytes is not None:
        st.download_button(
            "⬇️ Download Excel Template",
            data=xl_bytes,
            file_name="tokenomics_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.warning("Excel engines not found (install 'openpyxl' or 'xlsxwriter'). Download the CSV template instead.")

    st.download_button(
        "⬇️ Download CSV Template",
        data=csv_df.to_csv(index=False).encode("utf-8"),
        file_name="tokenomics_template.csv",
        mime="text/csv",
    )

with c2:
    uploaded = st.file_uploader("Upload your filled sheet (XLSX/CSV)", type=["xlsx", "xls", "csv"])
    use_project_info_from_sheet = st.checkbox(
        "Override Project Inputs from 'Project Info' sheet (if present)", value=True
    )

sheet_mode = False
sheet_allocations, sheet_tags, sheet_vesting = None, None, None

# -----------------------------
# Default Project Inputs (UI)
# -----------------------------
col1, col2, col3 = st.columns(3)
with col1:
    project_name = st.text_input("Project Name", value="MyProject")
    total_supply_tokens = st.number_input("Total Token Supply", min_value=1, value=1_000_000_000)
    tge_price = st.number_input("Token Price at TGE (USD)", min_value=0.00001, value=0.02, step=0.01, format="%.5f")
with col2:
    liquidity_fund = st.number_input("Liquidity Fund (USD)", min_value=0.0, value=500000.0, step=10000.0)
    project_type = st.selectbox("Project Category", ["Gaming", "DeFi", "NFT", "Infrastructure"], index=0)
    user_base = st.number_input("Current User Base", min_value=0, value=10000)
with col3:
    user_growth_rate = st.number_input("Monthly User Growth Rate (%)", min_value=0.0, value=10.0) / 100
    monthly_burn = st.number_input("Monthly Incentive Spend (USD)", min_value=0.0, value=50000.0, step=1000.0)
    revenue = st.number_input("Annual Revenue (USD)", min_value=0.0, value=1000000.0, step=10000.0)

# -----------------------------
# Upload parsing -> dicts
# -----------------------------
if uploaded:
    try:
        if uploaded.name.lower().endswith(".csv"):
            df_inputs = pd.read_csv(uploaded)
            df_proj = None
        else:
            try:
                df_inputs = pd.read_excel(uploaded, sheet_name="Tokenomics Inputs")
                try:
                    df_proj = pd.read_excel(uploaded, sheet_name="Project Info")
                except Exception:
                    df_proj = None
            except ImportError as ie:
                st.error("Reading .xlsx requires an Excel engine. Please `pip install openpyxl` (recommended) or upload a CSV using our template.")
                st.stop()

        df_inputs.columns = [str(c).strip() for c in df_inputs.columns]
        required_cols = {
            "Pool Name", "Allocation %", "Category", "TGE Unlock %",
            "Cliff (months)", "Vesting (months)", "Sellable at TGE (Yes/No)"
        }
        if not required_cols.issubset(set(df_inputs.columns)):
            st.error(f"Uploaded sheet is missing required columns:\n{sorted(list(required_cols - set(df_inputs.columns)))}")
            st.stop()

        def _to_float(x, default=0.0):
            try:
                return float(x)
            except Exception:
                return default

        def _to_int(x, default=0):
            try:
                return int(float(x))
            except Exception:
                return default

        sheet_allocations = {}
        sheet_tags = {}
        sheet_vesting = {}
        for _, row in df_inputs.iterrows():
            pool = str(row["Pool Name"]).strip()
            if not pool:
                continue
            alloc = _to_float(row["Allocation %"])
            tge = _to_float(row["TGE Unlock %"])
            cliff = _to_int(row["Cliff (months)"])
            vest = _to_int(row["Vesting (months)"])
            sellable_str = str(row["Sellable at TGE (Yes/No)"]).strip().lower()
            sellable = sellable_str in {"yes", "y", "true", "1"}

            sheet_allocations[pool] = alloc
            sheet_tags[pool] = str(row["Category"]).strip() if "Category" in row else "Other"
            sheet_vesting[pool] = {"cliff": cliff, "vesting": vest, "tge": tge, "sellable": sellable}

        total_from_sheet = sum(sheet_allocations.values())
        st.info(f"📊 Sheet parsed. Total allocation = **{total_from_sheet:.2f}%**.")
        if abs(total_from_sheet - 100.0) > 1e-6:
            st.warning("Allocation must total 100%. Please fix your sheet and re-upload.")
            st.stop()

        if use_project_info_from_sheet and df_proj is not None and len(df_proj) > 0:
            proj = df_proj.iloc[0].to_dict()
            project_name = str(proj.get("Project Name", project_name))
            total_supply_tokens = int(_to_float(proj.get("Total Supply", total_supply_tokens)))
            tge_price = _to_float(proj.get("TGE Price (USD)", tge_price))
            liquidity_fund = _to_float(proj.get("Liquidity Fund (USD)", liquidity_fund))
            project_type = str(proj.get("Project Category", project_type))
            user_base = int(_to_float(proj.get("Current User Base", user_base)))
            user_growth_rate = _to_float(proj.get("User Growth %", user_growth_rate * 100.0)) / 100.0
            monthly_burn = _to_float(proj.get("Monthly Incentive Spend (USD)", monthly_burn))
            revenue = _to_float(proj.get("Annual Revenue (USD)", revenue))
            st.success("✅ Project Info overridden from sheet.")
        sheet_mode = True

    except Exception as e:
        st.error(f"Could not parse uploaded file: {e}")
        st.stop()

# -----------------------------
# Step 1: Token Allocation & Vesting
# -----------------------------
st.header("📊 Step 1: Token Allocation & Vesting Schedule")

if not sheet_mode:
    st.markdown("Enter pool details in the table below:")
    pool_names = st.text_area("Pool Names (comma-separated)", value="Private Sale, Public Sale, Team, Ecosystem, Advisors, Listing, Airdrop, Staking", height=80)
    pools_list = [p.strip() for p in pool_names.split(",") if p.strip()]
    if len(pools_list) == 0:
        st.stop()

    # Manual table inputs
    col1m, col2m, col3m, col4m, col5m, col6m, col7m = st.columns(7)
    with col1m: st.markdown("**Pool Name**")
    with col2m: st.markdown("**Allocation %**")
    with col3m: st.markdown("**Category**")
    with col4m: st.markdown("**TGE Unlock %**")
    with col5m: st.markdown("**Cliff (months)**")
    with col6m: st.markdown("**Vesting (months)**")
    with col7m: st.markdown("**Sellable at TGE**")

    allocations, tags, vesting_schedule = {}, {}, {}
    total_alloc = 0.0

    for i, pool in enumerate(pools_list):
        c1r, c2r, c3r, c4r, c5r, c6r, c7r = st.columns(7)
        with c1r: st.markdown(f"**{pool}**")
        with c2r:
            allocations[pool] = st.number_input("", 0.0, 100.0, 0.0, step=1.0, key=f"alloc_{i}")
            total_alloc += allocations[pool]
        with c3r:
            tags[pool] = st.selectbox("", ["VC", "Community", "Team", "Liquidity", "Advisor", "Other"], key=f"tag_{i}")
        with c4r:
            tge = st.number_input("", 0.0, 100.0, 0.0, step=1.0, key=f"tge_{i}")
        with c5r:
            cliff = st.number_input("", 0, 48, 0, key=f"cliff_{i}")
        with c6r:
            vesting = st.number_input("", 0, 72, 12, key=f"vesting_{i}")
        with c7r:
            sellable_choice = st.selectbox("", ["Yes", "No"], key=f"sellable_{i}")
            sellable = (sellable_choice == "Yes")
        vesting_schedule[pool] = {"cliff": cliff, "vesting": vesting, "tge": tge, "sellable": sellable}

else:
    # Sheet-based inputs
    pools_list = list(sheet_allocations.keys())
    allocations = sheet_allocations
    tags = sheet_tags
    vesting_schedule = sheet_vesting
    total_alloc = sum(allocations.values())

    st.success("✅ Using inputs from uploaded sheet.")
    st.dataframe(pd.DataFrame([
        {
            "Pool": p,
            "Allocation %": allocations[p],
            "Category": tags.get(p, ""),
            "TGE % (of pool)": vesting_schedule[p]["tge"],
            "Cliff (m)": vesting_schedule[p]["cliff"],
            "Vesting (m)": vesting_schedule[p]["vesting"],
            "Sellable at TGE": "Yes" if vesting_schedule[p]["sellable"] else "No",
        } for p in pools_list
    ]))

# Validate total allocation
if total_alloc != 100:
    st.markdown(f"""
    <div class="warning-box">
        ⚠️ Allocation must total 100%. Current: {total_alloc:.2f}%
    </div>
    """, unsafe_allow_html=True)
    st.stop()
else:
    st.markdown(f"""
    <div class="success-box">
        ✅ Total allocation: {total_alloc:.0f}%
    </div>
    """, unsafe_allow_html=True)

#
# How to interpret the "Vesting (months)" column
vest_includes_cliff = st.checkbox(
    "Treat 'Vesting (months)' as TOTAL months including the cliff",
    value=False,
    help=(
        "ON → Linear months = Vesting − Cliff (common in spreadsheets).\n"
        "OFF → Linear months = Vesting (Cliff is separate)."
    ),
)
# -----------------------------
# Build vesting schedule dataframe (token-precise, cliff-safe)
# -----------------------------

def build_release_schedule(
    pools_list: list[str],
    allocations: dict,
    vesting_schedule: dict,
    total_supply_tokens: int,
    total_months: int = 72,
    catchup: bool = False,  # if True, unlocks the linear accrual that would have happened during the cliff at the cliff end
    vest_includes_cliff: bool = False,
):
    """Return a dataframe with Monthly Release Tokens/% and Cumulative columns.

    We compute in **tokens** first to avoid cumulative rounding drift and then
    derive percentages from tokens. Linear vesting starts strictly **after** the
    cliff at month `cliff + 1` unless `catchup=True`, in which case the linear
    accrual that would have occurred during the cliff is released at `month = cliff`.
    """
    months = list(range(total_months))
    df = pd.DataFrame({"Month": months})
    df["Monthly Release Tokens"] = 0.0

    for pool in pools_list:
        alloc_pct = float(allocations[pool])
        tge_pct_of_pool = float(vesting_schedule[pool].get("tge", 0.0))
        cliff = int(vesting_schedule[pool].get("cliff", 0))
        vest = int(vesting_schedule[pool].get("vesting", 0))

        # compute pool tokens
        pool_tokens = total_supply_tokens * (alloc_pct / 100.0)
        tge_tokens = pool_tokens * (tge_pct_of_pool / 100.0)
        linear_tokens_total = max(pool_tokens - tge_tokens, 0.0)

        release = [0.0] * total_months
        # TGE happens at Month 0
        release[0] += tge_tokens

        if vest > 0 and linear_tokens_total > 0:
            if catchup and cliff > 0:
                # unlock the linear accrual at the cliff end, then continue monthly
                accrued = linear_tokens_total * (cliff / vest)
                idx = min(cliff, total_months - 1)
                release[idx] += accrued
                remaining_months = max(vest - cliff, 1)
                per_month = (linear_tokens_total - accrued) / remaining_months
                start_m = cliff + 1
                end_m = cliff + vest
            else:
                # Determine linear months with a safety heuristic:
                # - If the sheet says "Vesting includes cliff" but vest <= cliff, assume the author
                #   actually entered a post‑cliff vest duration (common spreadsheet confusion).
                if vest_includes_cliff:
                    if vest <= cliff:
                        # Treat provided vest as post‑cliff months to avoid one‑month mega‑unlock
                        linear_months = max(vest, 1)
                    else:
                        linear_months = max(vest - cliff, 1)
                else:
                    linear_months = max(vest, 1)
                per_month = linear_tokens_total / linear_months
                start_m = cliff + 1
                end_m = cliff + linear_months

            for m in range(start_m, min(end_m, total_months - 1) + 1):
                release[m] += per_month

        # track per-pool and aggregate
        df[pool] = release
        df["Monthly Release Tokens"] += df[pool]

    # derive % from tokens (more stable numerically)
    df["Monthly Release %"] = (df["Monthly Release Tokens"] / float(total_supply_tokens)) * 100.0
    df["Cumulative Tokens"] = df["Monthly Release Tokens"].cumsum()
    df["Cumulative %"] = (df["Cumulative Tokens"] / float(total_supply_tokens)) * 100.0
    return df

# Build schedule using token-precise simulator (no cliff catch-up by default)
months = 72
df = build_release_schedule(
    pools_list=pools_list,
    allocations=allocations,
    vesting_schedule=vesting_schedule,
    total_supply_tokens=total_supply_tokens,
    total_months=months,
    catchup=False,  # set True if your vest design requires a cliff catch-up unlock
    vest_includes_cliff=vest_includes_cliff,
)

st.markdown("""
<div class="success-box">
    ✅ Vesting schedule processed successfully.
</div>
""", unsafe_allow_html=True)

# -----------------------------
# Metrics helpers
# -----------------------------
def inflation_guard(df: pd.DataFrame) -> list:
    inflation = []
    if "Cumulative Tokens" not in df.columns or len(df) == 0:
        return inflation
    cum = df["Cumulative Tokens"].values
    n = len(cum)
    for y in range(1, 7):
        if y == 1:
            start_idx = (y - 1) * 13  # 0
        else:
            start_idx = (y - 1) * 12
        if start_idx >= n:
            break
        end_idx = y * 12
        if end_idx >= n:
            end_idx = n - 1
        start_val = float(cum[start_idx])
        end_val = float(cum[end_idx])
        rate = ((end_val - start_val) / start_val) * 100.0 if start_val > 0 else 0.0
        inflation.append(round(rate, 2))
    return inflation

def calculate_monthly_supply_shock(df: pd.DataFrame) -> pd.Series:
    prev_circ = df["Cumulative Tokens"].shift(1).fillna(0.0)
    monthly_tokens = df["Monthly Release Tokens"]
    supply_shock = monthly_tokens.divide(prev_circ.replace(0, pd.NA)).fillna(0) * 100.0
    return supply_shock.round(2)

def shock_stopper(df: pd.DataFrame) -> dict:
    ss = calculate_monthly_supply_shock(df)
    ss = ss.iloc[1:] if len(ss) > 1 else ss  # exclude M0
    out = {
        "0–5%": int((ss <= 5).sum()),
        "5–10%": int(((ss > 5) & (ss <= 10)).sum()),
        "10–15%": int(((ss > 10) & (ss <= 15)).sum()),
        "15%+": int((ss > 15).sum()),
    }
    out[">10% months"] = out["10–15%"] + out["15%+"]
    out[">12% months"] = int((ss > 12).sum())
    out[">15% months"] = int((ss > 15).sum())
    out["% >10%"] = round(100 * out[">10% months"] / max(len(ss), 1), 2)
    return out

def governance_hhi(allocations):
    shares_decimal = [v / 100 for v in allocations.values()]
    hhi = sum([s ** 2 for s in shares_decimal])
    return round(hhi, 3)

def liquidity_shield_denominator(total_supply, tge_price, allocations, vesting, sellable_override=None):
    # If sellable_override provided as set of pool names, only include those;
    # otherwise include pools with sellable=True in vesting.
    sellable_pct = 0.0
    for p in allocations:
        if sellable_override is None:
            if vesting.get(p, {}).get("sellable"):
                sellable_pct += allocations[p] * (vesting[p]["tge"] / 100.0)
        else:
            if p in sellable_override:
                sellable_pct += allocations[p] * (vesting[p]["tge"] / 100.0)
    tokens_sellable = total_supply * (sellable_pct / 100.0)
    market_cap_at_tge = tokens_sellable * tge_price
    return sellable_pct, int(tokens_sellable), float(market_cap_at_tge)

def lockup_ratio(vesting, allocations, mode="supply"):
    eligible = [p for p in vesting if vesting[p].get("cliff", 0) >= 12]
    if mode == "pools":
        total = len(vesting) or 1
        return round(len(eligible) / total, 2)
    else:
        locked_percent = sum(allocations.get(p, 0.0) for p in eligible)
        return round(locked_percent / 100.0, 2)

def vc_dominance(allocations, tags):
    return round(sum([allocations[p] for p in allocations if tags.get(p) == "VC"]) / 100, 2)

def community_index(allocations, tags):
    return round(sum([allocations[p] for p in allocations if tags.get(p) == "Community"]) / 100, 2)

def emission_taper(df):
    first_12 = df.loc[0:11, "Monthly Release Tokens"].sum()
    last_12 = df["Monthly Release Tokens"].tail(12).sum()
    return round(first_12 / last_12 if last_12 > 0 else 0.0, 2)

def summarize_sim(sim_list):
    arr = np.array(sim_list, dtype=float)
    if arr.size == 0:
        return {"min": 0, "p25": 0, "median": 0, "p75": 0, "p90": 0, "max": 0}
    return {
        "min": float(np.min(arr)),
        "p25": float(np.percentile(arr, 25)),
        "median": float(np.median(arr)),
        "p75": float(np.percentile(arr, 75)),
        "p90": float(np.percentile(arr, 90)),
        "max": float(np.max(arr)),
    }

def monte_carlo_simulation(df, user_base, growth, months, category, total_supply, tge_price):
    # Simple survivability simulation with clamped moves and small noise
    if category == "Gaming":       arpu_monthly = (76 / 12) * 0.03
    elif category == "DeFi":       arpu_monthly = (3300 / 12) * 0.03
    elif category == "NFT":        arpu_monthly = (59 / 12) * 0.03
    else:                          arpu_monthly = (50 / 12) * 0.03

    rng = np.random.default_rng(seed=42)
    simulations = []
    for _ in range(100):
        price = float(tge_price)
        for m in range(min(months, len(df))):
            users = user_base * ((1 + growth) ** m)
            buy_pressure = users * arpu_monthly
            tokens_released = (df.loc[m, "Monthly Release %"] / 100.0) * total_supply
            sell_pressure = max(tokens_released * price, 1e-9)
            net_pressure = buy_pressure - sell_pressure
            price_change_factor = (net_pressure / sell_pressure)
            sensitivity = 0.10
            raw_move = price_change_factor * sensitivity
            clamped = float(np.clip(raw_move, -0.5, 2.0))  # −50% to +200% per month
            noise = rng.normal(0, 0.02)
            price *= max(0.0001, 1.0 + clamped + noise)
        simulations.append(price)
    return simulations

def game_theory_audit(hhi, shield, inflation):
    score = 0
    if hhi < 0.15: score += 1
    if shield >= 1.0: score += 1
    if len(inflation) > 0 and inflation[0] <= 300: score += 1
    if len(inflation) > 1 and inflation[1] <= 100: score += 1
    if len(inflation) > 2 and inflation[2] <= 50: score += 1
    label = "Excellent" if score >= 4 else "Moderate" if score == 3 else "Needs Improvement"
    return label, score

# -----------------------------
# Structured audit prompt helper
# -----------------------------
def build_structured_prompt(metrics: dict) -> str:
    import json as _json
    return f"""
ROLE
You are a senior tokenomics analyst. You audit token designs for investors and founders.
Be concise, structured, and institutional. Avoid fluff.
Do NOT invent numbers — only use values present inside <metrics>.

INPUT
<metrics>
{_json.dumps(metrics, indent=2)}
</metrics>

GLOBAL RULES
- Use ONLY metrics that exist in <metrics>. If a metric is missing, omit its section.
- Use the EXACT section titles below (important for rendering).
- For each section: include a one-line “Purpose — …” (definition) → “STAT:” line → “Price Impact — …” line → 2–3 “Suggestions:” bullets.
- Keep bullets short and action-oriented. Tie numbers → interpretation → price effect → fix.
- Values: keep units and significant figures; don’t round away meaning.
- Do not repeat definitions across metrics; keep each definition to one clear sentence.

OUTPUT STRUCTURE

1) Red Flags (first section)
- List the top 3–5 risks as single-line bullets in the format:
  - <Cause with metric + value> → <Effect on float/liquidity/governance> → <Impact on price/investor trust>.
- Choose the most material risks across all provided metrics.

2) Section by Section Analysis (use only metrics that exist in <metrics>)

YoY Inflation
Purpose — Define “YoY Inflation” in one sentence (change in circulating supply year over year).
STAT: Y1: <y1%>, Y2: <y2%>, Y3: <y3%>, Y4: <y4%>, Y5: <y5%>, Y6: <y6%>.
Price Impact — <One clear line on expected price/holder behavior given these %s>.
Suggestions:
- <Actionable fix #1 tied to mechanics/comms/sinks>
- <Actionable fix #2>
- <Actionable fix #3>

Supply Shock Bins
Purpose — Define “monthly supply shock” (monthly release as % of prior circulating) and why bins matter.
STAT: 0–5%: <n0_5> | 5–10%: <n5_10> | 10–15%: <n10_15> | 15%+: <n15p>; >10% months: <share>% — <diffuse/mixed/concentrated> release profile.
Price Impact — <One line on clustering near >10% months and expected drawdowns/liquidity needs>.
Suggestions:
- <Smooth or gate large unlocks via milestones/liquidity programs>
- <Stagger cliffs; convert to linear with caps>
- <Pre-announce market-making/liquidity buffers>

Governance HHI
Purpose — Define HHI as concentration index of token ownership/allocation (0–1, lower = more dispersed).
STAT: HHI: <hhi>.
Price Impact — <One line linking concentration to governance capture, sell pressure coordination, or signaling>.
Suggestions:
- <Broaden distribution, e.g., via community programs/lock-to-vote>
- <Cap single-pool allocations/introduce decay>
- <Strengthen quorum/anti-whale safeguards>

Liquidity Shield Ratio
Purpose — Define as Liquidity Fund / Sellable Mcap at TGE (higher = better defense).
STAT: Liquidity Shield: <ratio>x; Sellable at TGE: <sellable_pct>% (<sellable_tokens> tokens); Sellable Mcap: $<sellable_mcap>.
Price Impact — <One line on near-TGE volatility absorption and tail-risk>.
Suggestions:
- <Increase shield via larger liquidity fund/partner MM>
- <Reduce sellable at TGE/limit initial float>
- <Pair with fee rebates/LP incentives during volatile windows>

Lockup Ratio
Purpose — Define as share of supply/pools with cliff ≥12 months (stickier float).
STAT: Lockup (Supply): <lock_supply_pct>% | Lockup (Pools): <lock_pools_pct>% (eligibility: cliff ≥12m).
Price Impact — <One line on early-months float tightness vs delayed overhang>.
Suggestions:
- <Raise cliff on strategic pools / align with product milestones>
- <Introduce progressive vesting vs large cliffs>
- <Offer lock-to-earn programs for communities/teams>

VC Dominance
Purpose — Define as % of supply tagged “VC” (higher = potential sell/coordination risk).
STAT: VC Dominance: <vc_pct>%.
Price Impact — <One line on perceived overhang and investor signaling>.
Suggestions:
- <Redistribute via follow-on community rounds with lockups>
- <Hard commit to extended VC lock/earn-out>
- <Enhance disclosures on VC lock/participation>

Community Control Index
Purpose — Define as % of supply tagged “Community” (higher = alignment, usage flywheel).
STAT: Community Control: <community_pct>%.
Price Impact — <One line on adoption/retention vs speculation balance>.
Suggestions:
- <Route emissions to active users/revenue-linked rewards>
- <Delegate grants via on-chain KPIs>
- <Lock-to-use or usage-mining frameworks>

Emission Taper
Purpose — Define as ratio of tokens released in first 12m vs last 12m (front- vs back-loaded).
STAT: Emission Taper: <taper_ratio>x.
Price Impact — <One line on early sell pressure vs long-run sustainability>.
Suggestions:
- <Flatten early curve; move to milestone-gated linear>
- <Cap monthly max unlocks>
- <Tie emissions to real demand (users/revenue)>

Monte Carlo Survivability
Purpose — Define as distribution of simulated end-prices given buy/sell pressure model.
STAT: USD — min: <min>, p25: <p25>, median: <median>, p75: <p75>, p90: <p90>, max: <max>.
Price Impact — <One line on downside tails or upside skew and what drives it>.
Suggestions:
- <Reduce release rate in adverse months>
- <Boost demand levers (ARPU, active users) before major unlocks>
- <Add sinks/fees/redemptions>

Game Theory Score
Purpose — Define as composite score of structure (HHI, shield, inflation, etc.).
STAT: Score: <score>/5 — <label>.
Price Impact — <One line on overall investability/readiness for listings>.
Suggestions:
- <Target the weakest sub-metric from above>
- <Publish a clear unlock/liquidity and governance policy>
- <Stage reforms before major catalysts>

3) Final Summary (1–2 lines)
- One institutional line capturing float at TGE (<sellable_pct>% / $<sellable_mcap>), the standout strength, and the main risk.
- If FDV or other fields are NOT present in <metrics>, do not mention them.

STYLE GUIDANCE
- Bold only the section headers (exact titles above).
- Keep “Purpose — …”, “STAT: …”, “Price Impact — …”, then short “Suggestions:” bullets.
- Never add sections not listed. Never fabricate numbers. Keep it tight and decision-useful.
""".strip()

# -----------------------------
# Generate button (centered)
# -----------------------------
col_left, col_center, col_right = st.columns([4, 2, 4])
with col_center:
    generate = st.button("🚀 Generate Audit Report")

# -----------------------------
# Run audit
# -----------------------------
if generate:
    # Core metrics
    inflation = inflation_guard(df)
    df['Supply Shock %'] = calculate_monthly_supply_shock(df)
    shock = shock_stopper(df)
    hhi = governance_hhi(allocations)

    # Liquidity Shield — by default use 'sellable' flags from inputs
    sellable_pct, sellable_tokens, sellable_mc = liquidity_shield_denominator(
        total_supply_tokens, tge_price, allocations, vesting_schedule, sellable_override=None
    )
    shield = round(liquidity_fund / sellable_mc, 2) if sellable_mc > 0 else 0.0

    lock_supply = lockup_ratio(vesting_schedule, allocations, mode="supply")
    lock_pools = lockup_ratio(vesting_schedule, allocations, mode="pools")
    vc = vc_dominance(allocations, tags)
    community = community_index(allocations, tags)
    taper = emission_taper(df)
    monte = monte_carlo_simulation(df, user_base, user_growth_rate, 24, project_type, total_supply_tokens, tge_price)
    game_label, game_score = game_theory_audit(hhi, shield, inflation)

    # Score card
    st.markdown(f"""
    <div class="metric-card">
        <h3>🎯 Game Theory Audit Score: {game_score}/5</h3>
        <p style="font-size: 1.1rem; color: #FF6B35; margin: 0;">{game_label}</p>
    </div>
    """, unsafe_allow_html=True)

    # Year 1 table
    try:
        year1 = df.loc[0:12, ["Month", "Monthly Release Tokens", "Cumulative Tokens", "Cumulative %"]].copy()
        year1["Month"] = year1["Month"].apply(lambda m: f"M{int(m)}")
        year1["Monthly Release Tokens"] = year1["Monthly Release Tokens"].apply(lambda x: f"{x:,.0f}")
        year1["Cumulative Tokens"] = year1["Cumulative Tokens"].apply(lambda x: f"{x:,.0f}")
        year1["Cumulative %"] = year1["Cumulative %"].apply(lambda x: f"{x:.2f}%")
        st.markdown("### Year 1 Monthly Token Release (M0–M12)")
        st.table(year1)
        # Debug: by‑pool monthly breakdown (in millions)
        year1_by_pool = df.loc[0:12, ["Month"] + pools_list].copy()
        for p in pools_list:
            year1_by_pool[p] = (year1_by_pool[p] / 1_000_000).round(2)
        year1_by_pool["Month"] = year1_by_pool["Month"].apply(lambda m: f"M{int(m)}")
        with st.expander("🔎 Year 1 breakdown by pool (millions)", expanded=False):
            st.dataframe(year1_by_pool)
    except Exception as e:
        st.markdown(f"<div class='warning-box'>Could not render Year 1 table: {e}</div>", unsafe_allow_html=True)

    # Build figures (do not display here; will place under matching sections)
    # TDeFi theme colors
    tdefi_yellow = "#ffb703"  # warm yellow
    tdefi_orange = "#fd7e14"

    fig1, ax1 = plt.subplots(figsize=(6, 4))
    years = list(range(1, len(inflation) + 1))
    bars = ax1.bar(years, inflation, color=tdefi_yellow, alpha=0.95, edgecolor="#333333")
    ax1.set_title("Inflation Guard", color="#000000")
    ax1.set_xlabel("Year", color="#000000"); ax1.set_ylabel("%", color="#000000")
    ax1.grid(True, axis='y', alpha=0.25)
    # Set y-limits with headroom so labels don't collide with top
    max_val = max(inflation) if len(inflation) else 0
    ax1.set_ylim(0, max_val * 1.18 + (5 if max_val < 50 else 0))
    for rect, val in zip(bars, inflation):
        ax1.text(rect.get_x() + rect.get_width()/2.0, rect.get_height() * 1.02,
                 f"{val:.0f}%", ha='center', va='bottom', fontsize=9, color="#000000")
    ax1.tick_params(colors="#000000")

    fig2, ax2 = plt.subplots(figsize=(6, 4))
    ordered_bins = ["0–5%", "5–10%", "10–15%", "15%+"]
    ax2.bar(ordered_bins, [shock[b] for b in ordered_bins], color=tdefi_orange, alpha=0.95, edgecolor="#333333")
    ax2.set_title("Shock Stopper", color="#000000")
    ax2.grid(True, axis='y', alpha=0.25)
    ax2.tick_params(colors="#000000")
    ax2.set_ylim(0, max([shock[b] for b in ordered_bins]) * 1.18 + 1)

    fig3, ax3 = plt.subplots(figsize=(6, 4))
    if len(set(np.round(monte, 6))) > 1:
        ax3.hist(monte, bins=min(20, len(set(np.round(monte, 6)))), color=tdefi_yellow, edgecolor="#333333")
        ax3.set_title("Monte Carlo Survivability", color="#000000")
    else:
        ax3.bar(["Simulated"], [monte[0]], color=tdefi_yellow, edgecolor="#333333")
        ax3.set_title("Simulation Output", color="#000000")
    ax3.grid(True, axis='y', alpha=0.25)
    ax3.tick_params(colors="#000000")

    # Optional: supply shock early months figure (kept for future use, not auto-rendered)
    fig_shock = None
    try:
        shocks = df.loc[1:5, ['Month', 'Supply Shock %']].copy()
        fig_shock, ax_shock = plt.subplots(figsize=(10, 4))
        x_labels = [f"M{int(m)}" for m in shocks['Month']]
        values = shocks['Supply Shock %'].tolist()
        ax_shock.bar(x_labels, values)
        ax_shock.set_title("Major Supply Shocks (Early Months)")
        ax_shock.set_ylabel("%")
        for i, v in enumerate(values):
            ax_shock.text(i, v * 1.01, f"{v:.2f}%", ha='center', va='bottom', fontsize=9)
    except Exception:
        fig_shock = None

    # -----------------------------
    # Grounded AI analysis
    # -----------------------------
    api_key = get_openai_api_key()
    if not api_key:
        st.markdown("""
        <div class="warning-box">
            ❌ Missing OpenAI API key. Set OPENAI_API_KEY env var or add st.secrets['openai']['api_key'].
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    client = OpenAI(api_key=api_key)

    metrics = {
        "project_name": project_name,
        "total_supply": int(total_supply_tokens),
        "tge_price_usd": float(tge_price),
        "liquidity_fund_usd": float(liquidity_fund),
        "sellable_at_tge_pct": float(round(sellable_pct, 4)),
        "sellable_at_tge_tokens": int(sellable_tokens),
        "sellable_mcap_at_tge_usd": float(round(sellable_mc, 2)),
        "yoy_inflation_pct": inflation,
        "shock_bins": shock,
        "governance_hhi": float(hhi),
        "liquidity_shield_ratio": float(shield),
        "lockup_ratio_supply_pct": float(round(lock_supply * 100, 2)),
        "lockup_ratio_pools_pct": float(round(lock_pools * 100, 2)),
        "vc_dominance_pct": float(round(vc * 100, 2)),
        "community_index_pct": float(round(community * 100, 2)),
        "emission_taper_ratio": float(taper),
        "monte_carlo_summary_price_usd": summarize_sim(monte),
        "game_theory_score": int(game_score),
        "game_theory_label": game_label,
    }

    with st.spinner("🤖 AI is analyzing your tokenomics..."):
        analysis_prompt = build_structured_prompt(metrics)

        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are a senior tokenomics audit analyst."},
                {"role": "user", "content": analysis_prompt}
            ],
            temperature=0.3
        )
        summary = response.choices[0].message.content

    # Parse and render AI analysis into aligned sections with graphs placed contextually
    st.markdown("""
    <div class="metric-card">
        <h3>🤖 AI Tokenomics Analysis</h3>
    </div>
    """, unsafe_allow_html=True)

    SECTION_TITLES = [
        "Red Flags",
        "Section by Section Analysis",
        "YoY Inflation",
        "Supply Shock Bins",
        "Supply Shock",
        "Supply Shock bins",
        "Governance HHI",
        "Liquidity Shield Ratio",
        "Lockup Ratio",
        "VC Dominance",
        "Community Control Index",
        "Emission Taper",
        "Monte Carlo Survivability",
        "Game Theory Score",
    ]

    def _strip_leading_bullets(s: str) -> str:
        return re.sub(r'^[\s\-\*•·]+', '', s or '').strip()

    def _split_title_subtitle(s: str, section_names: list[str]):
        raw = s.strip()
        # Normalize: drop leading numbering/emojis/bullets like "1) ", "🚩 ", "- "
        raw = re.sub(r'^(?:\d+\)|\d+\.|[\-\*•·]+|🚩|🔴|🟠|🟡|🔵|🔒|💼|👥|📉|🎲|🧠)\s*', '', raw)
        for name in section_names:
            if raw.lower().startswith(name.lower()):
                rest = raw[len(name):].lstrip()
                rest = re.sub(r'^(—|–|-|:)\s*', '', rest)
                rest = re.sub(r'^\s{2,}', '', rest)
                return (name, rest)
        return None

    def parse_ai_summary_sections(text: str):
        lines = (text or '').replace('\r\n', '\n').replace('\r', '\n').split('\n')
        # Find header indices
        headers = []  # (idx, title, subtitle)
        for i, raw in enumerate(lines):
            nob = _strip_leading_bullets(raw)
            split = _split_title_subtitle(nob, SECTION_TITLES)
            if split:
                headers.append((i, split[0], split[1]))
        # If none detected, return one generic section
        if not headers:
            return [{
                'title': 'Summary', 'subtitle': '', 'stat': '', 'price': '', 'explainer': [], 'suggestions': []
            }]
        sections = []
        for h_idx, (idx, title, subtitle) in enumerate(headers):
            end = headers[h_idx + 1][0] if h_idx + 1 < len(headers) else len(lines)
            chunk = [l.strip() for l in lines[idx+1:end] if l.strip()]
            purpose = ''
            stat = ''
            price_line = ''
            tail_bullets = []
            expl_paras = []
            for ln in chunk:
                nob = _strip_leading_bullets(ln)
                # Purpose
                if re.match(r'^purpose\b', nob, flags=re.I):
                    purpose = re.sub(r'^purpose\s*[—:-]?\s*', '', nob, flags=re.I)
                    continue
                # STAT (accept both 'STAT' and 'Impact' labels)
                if nob.upper().startswith('STAT') or nob.upper().startswith('IMPACT'):
                    stat = re.sub(r"^(STAT|Impact)\s*(Impact)?\s*[:\-—]?\s*", '', nob, flags=re.I)
                    continue
                # Price/Investor (or Price Impact) — treat as a single bullet line
                if re.match(r'^(price\s*impact|price/investor)\b', nob, flags=re.I):
                    price_line = re.sub(r'^(price\s*impact|price/investor)\s*[:\-—]?\s*', '', nob, flags=re.I).strip()
                    continue
                # Ignore explicit 'Suggestions:' label lines per new prompt
                if re.match(r'^suggestions?\s*:\s*$', nob, flags=re.I):
                    continue
                # Bullet → suggestion
                if re.match(r'^([\-\*•·]+)\s+.+$', ln):
                    s = _strip_leading_bullets(ln)
                    tail_bullets.append(s)
                    continue
                # Otherwise explainer paragraph
                # Ignore any 'Here ... means -' lines to keep the layout minimal
                if re.match(r'^here\b', nob, flags=re.I):
                    continue
                expl_paras.append(nob)
            sections.append({
                'title': title,
                'subtitle': subtitle,
                'purpose': purpose,
                'stat': stat,
                'price': price_line,
                'tail_bullets': tail_bullets,
                'explainer': expl_paras,
            })
        return sections

    sections = parse_ai_summary_sections(summary)

    # Remove duplicate 'Game Theory Score' sections, keep only the last one
    def _dedupe_keep_last(sections_list, title_name):
        last_index = None
        for idx, sec in enumerate(sections_list):
            if (sec.get('title') or '').strip().lower() == title_name.lower():
                last_index = idx
        if last_index is None:
            return sections_list
        out = []
        for idx, sec in enumerate(sections_list):
            if (sec.get('title') or '').strip().lower() == title_name.lower() and idx != last_index:
                continue
            out.append(sec)
        return out

    sections = _dedupe_keep_last(sections, 'Game Theory Score')

    # Map figures to sections
    fig_map = {
        'yoy inflation': fig1,
        'supply shock': fig2,
        'shock stopper': fig2,
        'monte carlo survivability': fig3,
        'monte carlo': fig3,
    }

    for sec in sections:
        title = sec['title']
        subtitle = sec.get('subtitle') or ''
        purpose = sec.get('purpose') or ''
        stat = sec.get('stat') or ''
        price_text = sec.get('price') or ''
        tail_bullets = sec.get('tail_bullets') or []
        expl = sec.get('explainer') or []

        st.markdown(f"<div class='analysis-section'>", unsafe_allow_html=True)
        # Distinct style for Red Flags
        key_title = title.lower().strip()
        if key_title == 'red flags':
            st.markdown(f"<div class='analysis-title' style='color:#ff4d4f;'>🚩 {title}</div>", unsafe_allow_html=True)
        elif key_title == 'section by section analysis':
            st.markdown(f"<div class='analysis-title' style='font-size:1.3rem;'>{title}</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div class='analysis-title'>{title}</div>", unsafe_allow_html=True)
        # Purpose (prefer explicit Purpose; fallback to subtitle)
        if purpose:
            st.markdown(f"<div class='analysis-body'><b>Purpose -</b> {purpose}</div>", unsafe_allow_html=True)
        elif subtitle:
            st.markdown(f"<div class='analysis-body'>{subtitle}</div>", unsafe_allow_html=True)
        if stat:
            st.markdown(f"<div class='analysis-stat'>{stat}</div>", unsafe_allow_html=True)

        # Place figure immediately after STAT
        key = title.lower().strip()
        if 'supply shock' in key:
            key = 'supply shock'
        if 'yoy' in key or 'year-over-year' in key:
            key = 'yoy inflation'
        if 'monte carlo' in key:
            key = 'monte carlo survivability'
        fig_to_show = fig_map.get(key)
        if stat and fig_to_show is not None and title.lower().strip() != 'red flags':
            st.pyplot(fig_to_show, clear_figure=False)

        # Price Impact bullet then remaining bullets
        if price_text:
            st.markdown(f"- Price Impact — {price_text}", unsafe_allow_html=True)
        if tail_bullets:
            st.markdown("<b>Suggestions -</b>", unsafe_allow_html=True)
        for item in tail_bullets:
            st.markdown(f"- {item}")

        st.markdown("</div>", unsafe_allow_html=True)
        # Divider between metrics
        st.markdown("---")

    # -----------------------------
    # PDF Report
    # -----------------------------
    def strip_md(text: str) -> str:
        """Remove basic markdown like **bold** and inline code."""
        t = text or ""
        t = re.sub(r"\*\*(.+?)\*\*", r"\1", t)
        t = re.sub(r"`([^`]*)`", r"\1", t)
        t = t.replace("**", "").replace("*", "") 
        return t.strip()

    def _find_logo_path():
        """Try common filenames for a logo to place in the PDF header."""
        for fn in [
            "your_logo.jpeg",
        ]:
            if os.path.exists(fn):
                return fn
        return None

    # Check if we can render unicode (emojis) by using a local TTF font if present.
    UNICODE_FONT = None
    for ttf in ["DejaVuSans.ttf", "NotoSans-Regular.ttf"]:
        if os.path.exists(ttf):
            UNICODE_FONT = ttf
            break
    # Optional bold unicode font (for true bold headings if available)
    UNICODE_FONT_BOLD = None
    for ttf in ["DejaVuSans-Bold.ttf", "NotoSans-Bold.ttf"]:
        if os.path.exists(ttf):
            UNICODE_FONT_BOLD = ttf
            break

    def dash_safe(text: str) -> str:
        """Replace Unicode dashes with ASCII hyphen-minus for core fonts."""
        return (text or "").replace("—", "-").replace("–", "-")

    def title_with_emoji(title: str) -> str:
        """Return the plain title (no emojis) to match the desired format."""
        return title.strip()

    class PDF(FPDF):
        def __init__(self, logo_path=None):
            super().__init__()
            self.logo_path = logo_path

        def header(self):
            # Only show the logo at the top-right, scaled to 60%
            if self.logo_path and os.path.exists(self.logo_path):
                try:
                    logo_w = 13  # ~60% of previous 22mm
                    x_pos = self.w - self.r_margin - logo_w
                    self.image(self.logo_path, x=x_pos, y=8, w=logo_w)
                except Exception:
                    pass
            # First page: centered title
            if self.page_no() == 1:
                try:
                    if UNICODE_FONT:
                        self.add_font("DejaVu", "", UNICODE_FONT, uni=True)
                        if UNICODE_FONT_BOLD:
                            self.add_font("DejaVu", "B", UNICODE_FONT_BOLD, uni=True)
                        self.set_font("DejaVu", "", 16)
                    else:
                        self.set_font("Arial", "B", 16)
                except Exception:
                    self.set_font("Arial", "B", 16)
                self.set_y(10)
                self.cell(0, 10, sanitize_text("Tokenomics Audit Report"), ln=True, align="C")
                # Optional small project name under title
                try:
                    if UNICODE_FONT:
                        self.set_font("DejaVu", "", 11)
                    else:
                        self.set_font("Arial", "I", 11)
                except Exception:
                    self.set_font("Arial", "I", 11)
                self.cell(0, 7, sanitize_text(project_name), ln=True, align="C")
                self.ln(2)
            else:
                # Small spacer to separate header from body on subsequent pages
                self.ln(8)

        def footer(self):
            self.set_y(-15)
            self.set_font("Arial", "I", 8)
            self.cell(0, 10, "Report by TDeFi", 0, 0, "C")

    # Sanitize helper aware of unicode capability
    def sanitize_text(text):
        t = (text or "").strip()
        if UNICODE_FONT:
            return t  # keep unicode (emojis)
        # Replace unicode dashes with ASCII hyphen before latin-1 encoding to avoid dropping ranges like 0–5%
        t = dash_safe(t)
        # ensure latin-1 safe for default Arial
        return t.encode("latin-1", "ignore").decode("latin-1")

    def normalize_ai_summary(text: str) -> str:
        """
        Insert smart breaks so 'Price Impact —' (or legacy 'Price/Investor —')
        starts on its own line, and tidy excessive spaces.
        """
        if not text:
            return ""
        t = text.replace("\r\n", "\n").replace("\r", "\n")
        # Break after STAT sentences when 'Price Impact —' or 'Price/Investor —' follows on the same line
        t = re.sub(r"\.\s+Price\s*Impact\s+—\s+", ".\nPrice Impact — ", t, flags=re.I)
        t = re.sub(r"\.\s+Price/Investor\s+—\s+", ".\nPrice/Investor — ", t, flags=re.I)
        # Collapse long spaces
        t = re.sub(r"[ \t]+", " ", t)
        return t.strip()

    def render_structured_summary(pdf: FPDF, summary_text: str, effective_page_width: float, base_font_size: int = 11, section_images: Optional[dict] = None):
        """
        Renders the AI summary preserving structure:
        - Headings like 'YoY Inflation — ...' become bold titles (with emojis if available) + italic subtitle.
        - Lines beginning with 'Price Impact —' (or 'Price/Investor —' legacy) are emphasized.
        - Lines starting with '- ' are bullets (kept).
        - Remaining lines are paragraphs.
        """
        BULLET = "•" if UNICODE_FONT else "\xb7"  # prefer solid bullet when unicode font available
        SECTION_SPACING = 2
        line_h = 7
        BULLET_CELL_W = 3.5
        INDENT = 6

        def _insert_section_image_for(title: str):
            if not section_images or not title:
                return False
            key = title.lower()
            if 'supply shock' in key:
                key_norm = 'supply shock'
            elif 'yoy' in key or 'inflation' in key:
                key_norm = 'yoy inflation'
            elif 'monte carlo' in key:
                key_norm = 'monte carlo'
            else:
                key_norm = key
            path = None
            if key_norm in section_images:
                path = section_images[key_norm]
            else:
                for k, v in section_images.items():
                    if k in key_norm:
                        path = v; break
            if path and os.path.exists(path):
                pdf.ln(1)
                fig_width = effective_page_width
                pdf.image(path, x=pdf.l_margin, w=fig_width)
                pdf.ln(2)
                return True
            return False

        def _strip_leading_bullets(s: str) -> str:
            # Remove any leading bullet markers and surrounding spaces: -, *, •, ·
            return re.sub(r'^[\s\-\*•·]+', '', s or '').strip()

        def _split_title_subtitle(s: str, section_names: list[str]) -> tuple[str, str] | None:
            """
            Try to split a line into (title, subtitle) when it starts with a known section name.
            Accept separators: em/en dash, hyphen, colon, or 2+ spaces.
            Returns (title, subtitle) or None if not a section header.
            """
            raw = s.strip()
            # If raw begins with a known title, try multiple separators
            for name in section_names:
                if raw.lower().startswith(name.lower()):
                    rest = raw[len(name):].lstrip()
                    if not rest:
                        return (name, "")
                    # Strip common separators
                    rest = re.sub(r'^(—|–|-|:)\s*', '', rest)
                    # If still begins with multiple spaces, treat as separator
                    rest = re.sub(r'^\s{2,}', '', rest)
                    return (name, rest)
            # If it contains a dash separator like "Title — subtitle"
            m = re.match(r'^(.+?)\s*[—–-]\s*(.+)$', raw)
            if m and any(raw.lower().startswith(n.lower()) for n in section_names):
                return (m.group(1).strip(), m.group(2).strip())
            return None

        # Optionally switch to a Unicode font so emojis render
        if UNICODE_FONT:
            try:
                pdf.add_font("DejaVu", "", UNICODE_FONT, uni=True)
                pdf.set_font("DejaVu", "", base_font_size)
            except Exception:
                pdf.set_font("Arial", "", base_font_size)
        else:
            pdf.set_font("Arial", "", base_font_size)

        lines = normalize_ai_summary(summary_text).splitlines()
        SECTION_TITLES = [
            "Red Flags",
            "Section by Section Analysis",
            "YoY Inflation",
            "Supply Shock Bins",
            "Supply Shock",
            "Supply Shock bins",
            "Governance HHI",
            "Liquidity Shield Ratio",
            "Lockup Ratio",
            "VC Dominance",
            "Community Control Index",
            "Emission Taper",
            "Monte Carlo Survivability",
            "Game Theory Score",
        ]
        current_section = None
        inserted_image_for_current = False
        for raw in lines:
            line = raw.strip()
            if not line:
                continue

            # Skip definition helper lines like 'Here ... means -'
            if re.match(r'^here\b', line, flags=re.I):
                continue

            # Try to parse known section headers first (works for both "· YoY Inflation  ..." and "YoY Inflation — ...")
            line_nobullet = _strip_leading_bullets(line)
            split = _split_title_subtitle(line_nobullet, SECTION_TITLES)
            if split is not None:
                title, subtitle = split
                # New section starts
                current_section = title
                inserted_image_for_current = False
                pdf.ln(2)
                # Render section title larger & bold; Heading 1 for top sections
                is_redflags = (title.strip().lower() == 'red flags')
                is_section2 = (title.strip().lower() == 'section by section analysis')
                if is_redflags:
                    pdf.set_text_color(200, 0, 0)
                else:
                    pdf.set_text_color(0, 0, 0)
                if UNICODE_FONT:
                    pdf.set_font("DejaVu", "", (base_font_size + 6) if (is_redflags or is_section2) else (base_font_size + 4))
                else:
                    pdf.set_font("Arial", "B", (base_font_size + 5) if (is_redflags or is_section2) else (base_font_size + 3))
                pdf.multi_cell(effective_page_width, line_h + 3, sanitize_text(title_with_emoji(strip_md(title))))
                # Subtitle (smaller / italic style)
                if subtitle:
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size + 0)
                    else:
                        pdf.set_font("Arial", "I", base_font_size)
                    pdf.multi_cell(effective_page_width, line_h, sanitize_text(strip_md(subtitle)))
                # Reset body font
                pdf.set_text_color(0, 0, 0)
                if UNICODE_FONT:
                    pdf.set_font("DejaVu", "", base_font_size)
                else:
                    pdf.set_font("Arial", "", base_font_size)
                # Hard reset X to left margin to avoid any drift before section body
                pdf.set_x(pdf.l_margin)
                pdf.ln(1)
                continue

            # Lines starting with '## ' should render bold without the '##'
            m_hash = re.match(r'^##\s*(.+)$', line)
            if m_hash:
                text = strip_md(m_hash.group(1))
                if UNICODE_FONT and 'DejaVu' in (pdf.font_family or '') and 'B' in (pdf.font_style or ''):
                    # already bold, just render
                    pass
                try:
                    if UNICODE_FONT and UNICODE_FONT_BOLD:
                        pdf.set_font("DejaVu", "B", base_font_size)
                    elif UNICODE_FONT:
                        # Emulate bold with a slightly larger size
                        pdf.set_font("DejaVu", "", base_font_size + 1)
                    else:
                        pdf.set_font("Arial", "B", base_font_size)
                except Exception:
                    pdf.set_font("Arial", "B", base_font_size)
                pdf.multi_cell(effective_page_width, line_h, sanitize_text(text))
                # Reset to body font
                if UNICODE_FONT:
                    pdf.set_font("DejaVu", "", base_font_size)
                else:
                    pdf.set_font("Arial", "", base_font_size)
                pdf.ln(0.5)
                continue

            # If a line is a STAT/Impact line, render text (without the label); insert figure right after
            if re.match(r"^(STAT|Impact)\b", line, flags=re.I):
                if UNICODE_FONT:
                    pdf.set_font("DejaVu", "", base_font_size)
                else:
                    pdf.set_font("Arial", "B", base_font_size)
                left_margin = pdf.l_margin + INDENT
                cur_y = pdf.get_y()
                pdf.set_xy(left_margin, cur_y)
                # Cleanup: remove 'STAT'/'STAT Impact' prefix and expand 'Y1–Y6:' pattern
                clean = re.sub(r"^(STAT|Impact)\s*(Impact)?\s*[:\-—]?\s*", "", line, flags=re.I)
                clean = re.sub(r"Y1\s*[–-]?\s*Y6\s*:\s*", "Y1, Y2, Y3, Y4, Y5, Y6: ", clean)
                clean = re.sub(r"Y1Y6\s*:\s*", "Y1, Y2, Y3, Y4, Y5, Y6: ", clean)
                # For YoY Inflation, bold the label prefix 'Y1, Y2, Y3, Y4, Y5, Y6:' and keep the rest normal
                did_split = False
                if current_section and 'yoy' in current_section.lower():
                    m = re.match(r"^(\s*Y\s*1\s*,\s*Y\s*2\s*,\s*Y\s*3\s*,\s*Y\s*4\s*,\s*Y\s*5\s*,\s*Y\s*6\s*:)\s*(.*)$", strip_md(clean), flags=re.I)
                    if m:
                        label_txt = m.group(1).strip() + ' '
                        body_txt = m.group(2).strip()
                        # Bold label
                        try:
                            if UNICODE_FONT and UNICODE_FONT_BOLD:
                                pdf.set_font("DejaVu", "B", base_font_size)
                            elif UNICODE_FONT:
                                pdf.set_font("DejaVu", "", base_font_size + 1)
                            else:
                                pdf.set_font("Arial", "B", base_font_size)
                        except Exception:
                            pdf.set_font("Arial", "B", base_font_size)
                        pdf.cell(pdf.get_string_width(sanitize_text(label_txt)) + 1, line_h, sanitize_text(label_txt), ln=0)
                        # Normal body
                        if UNICODE_FONT:
                            pdf.set_font("DejaVu", "", base_font_size)
                        else:
                            pdf.set_font("Arial", "", base_font_size)
                        start_x = pdf.get_x()
                        used_width = start_x - (pdf.l_margin + INDENT)
                        pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(body_txt))
                        did_split = True
                if not did_split:
                    # Render full line in bold (fallback)
                    try:
                        if UNICODE_FONT and UNICODE_FONT_BOLD:
                            pdf.set_font("DejaVu", "B", base_font_size)
                        elif UNICODE_FONT:
                            pdf.set_font("DejaVu", "", base_font_size + 1)
                        else:
                            pdf.set_font("Arial", "B", base_font_size)
                    except Exception:
                        pdf.set_font("Arial", "B", base_font_size)
                    pdf.multi_cell(effective_page_width - INDENT, line_h, sanitize_text(strip_md(clean)))
                pdf.set_xy(pdf.l_margin, pdf.get_y())
                pdf.ln(0.5)
                # Insert the figure once, immediately after STAT line
                if not inserted_image_for_current:
                    inserted_image_for_current = _insert_section_image_for(current_section)
                continue

            # Purpose (non-bullet) — bold label + normal body
            if re.match(r'^purpose\b', line, flags=re.I):
                left_margin = pdf.l_margin + INDENT
                pdf.set_xy(left_margin, pdf.get_y())
                body = re.sub(r'^purpose\s*[—:-]?\s*', '', line, flags=re.I)
                try:
                    if UNICODE_FONT and UNICODE_FONT_BOLD:
                        pdf.set_font("DejaVu", "B", base_font_size)
                    elif UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size + 1)
                    else:
                        pdf.set_font("Arial", "B", base_font_size)
                except Exception:
                    pdf.set_font("Arial", "B", base_font_size)
                label = "Purpose — "
                pdf.cell(pdf.get_string_width(sanitize_text(label)) + 1, line_h, sanitize_text(label), ln=0)
                if UNICODE_FONT:
                    pdf.set_font("DejaVu", "", base_font_size)
                else:
                    pdf.set_font("Arial", "", base_font_size)
                start_x = pdf.get_x()
                used_width = start_x - (pdf.l_margin + INDENT)
                pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(strip_md(body)))
                pdf.set_xy(pdf.l_margin, pdf.get_y())
                pdf.ln(0.5)
                continue

            # Suggestions header (non-bullet) — bold line
            if re.match(r'^suggestions?\s*[:\-—]?\s*$', line, flags=re.I):
                left_margin = pdf.l_margin + INDENT
                pdf.set_xy(left_margin, pdf.get_y())
                try:
                    if UNICODE_FONT and UNICODE_FONT_BOLD:
                        pdf.set_font("DejaVu", "B", base_font_size)
                    elif UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size + 1)
                    else:
                        pdf.set_font("Arial", "B", base_font_size)
                except Exception:
                    pdf.set_font("Arial", "B", base_font_size)
                pdf.multi_cell(effective_page_width - INDENT, line_h, sanitize_text("Suggestions —"))
                if UNICODE_FONT:
                    pdf.set_font("DejaVu", "", base_font_size)
                else:
                    pdf.set_font("Arial", "", base_font_size)
                pdf.ln(0.5)
                continue

            # Price Impact bullet
            if re.match(r"^(price\s*impact)\b", line, flags=re.I):
                # Render as bullet list with a 'Price Impact —' label prefix
                left_margin = pdf.l_margin + INDENT
                cur_y = pdf.get_y()
                pdf.set_xy(left_margin, cur_y)
                pdf.cell(BULLET_CELL_W, line_h, BULLET, ln=0)
                label = "Price Impact — "
                if UNICODE_FONT:
                    # DejaVu regular only; simulate emphasis by regular
                    pdf.set_font("DejaVu", "", base_font_size)
                else:
                    pdf.set_font("Arial", "B", base_font_size)
                safe_label = sanitize_text(label)
                pdf.cell(pdf.get_string_width(safe_label) + 1, line_h, safe_label, ln=0)
                body = re.sub(r"^(price\s*impact)\s*[—:-]?\s*", "", line, flags=re.I)
                if UNICODE_FONT:
                    pdf.set_font("DejaVu", "", base_font_size)
                else:
                    pdf.set_font("Arial", "", base_font_size)
                start_x = pdf.get_x()
                used_width = start_x - (pdf.l_margin + INDENT)
                pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(strip_md(body)))
                pdf.set_xy(pdf.l_margin, pdf.get_y())
                pdf.ln(0.5)
                continue

            # Bulleted lines
            m_bullet = re.match(r'^([\-\*•·]+)\s+(.*)$', raw)
            if m_bullet:
                content = m_bullet.group(2).strip()
                # Suggestions header (bullet) — bold line
                if re.match(r'^suggestions?\s*[:\-—]?\s*$', content, flags=re.I):
                    left_margin = pdf.l_margin + INDENT
                    pdf.set_xy(left_margin, pdf.get_y())
                    try:
                        if UNICODE_FONT and UNICODE_FONT_BOLD:
                            pdf.set_font("DejaVu", "B", base_font_size)
                        elif UNICODE_FONT:
                            pdf.set_font("DejaVu", "", base_font_size + 1)
                        else:
                            pdf.set_font("Arial", "B", base_font_size)
                    except Exception:
                        pdf.set_font("Arial", "B", base_font_size)
                    pdf.multi_cell(effective_page_width - INDENT, line_h, sanitize_text("Suggestions —"))
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size)
                    else:
                        pdf.set_font("Arial", "", base_font_size)
                    pdf.set_xy(pdf.l_margin, pdf.get_y())
                    pdf.ln(0.5)
                    continue
                # Skip definition bullets (Front-loaded / Back-loaded, Concentrated, etc.)
                if re.match(r'^(front\-?loaded|back\-?loaded|moderate|concentrated|diffuse|mixed|low|high|below|at|above|tight|loose|elevated|modest|strong|weak|balanced)\s*:', content, flags=re.I):
                    continue
                # Bullet line that begins with '## ' → bold content without the hashes
                m_b_content = re.match(r'^##\s*(.+)$', content)
                if m_b_content:
                    btxt = strip_md(m_b_content.group(1))
                    left_margin = pdf.l_margin + INDENT
                    cur_y = pdf.get_y()
                    pdf.set_xy(left_margin, cur_y)
                    pdf.cell(BULLET_CELL_W, line_h, BULLET, ln=0)
                    try:
                        if UNICODE_FONT and UNICODE_FONT_BOLD:
                            pdf.set_font("DejaVu", "B", base_font_size)
                        elif UNICODE_FONT:
                            pdf.set_font("DejaVu", "", base_font_size + 1)
                        else:
                            pdf.set_font("Arial", "B", base_font_size)
                    except Exception:
                        pdf.set_font("Arial", "B", base_font_size)
                    start_x = pdf.get_x()
                    used_width = start_x - (pdf.l_margin + INDENT)
                    pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(btxt))
                    # Reset font to body
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size)
                    else:
                        pdf.set_font("Arial", "", base_font_size)
                    pdf.set_xy(pdf.l_margin, pdf.get_y())
                    pdf.ln(0.5)
                    continue
                # Purpose bullet — bold label + normal body
                if re.match(r'^purpose\b', content, flags=re.I):
                    left_margin = pdf.l_margin + INDENT
                    pdf.set_xy(left_margin, pdf.get_y())
                    body = re.sub(r'^purpose\s*[—:-]?\s*', '', content, flags=re.I)
                    try:
                        if UNICODE_FONT and UNICODE_FONT_BOLD:
                            pdf.set_font("DejaVu", "B", base_font_size)
                        elif UNICODE_FONT:
                            pdf.set_font("DejaVu", "", base_font_size + 1)
                        else:
                            pdf.set_font("Arial", "B", base_font_size)
                    except Exception:
                        pdf.set_font("Arial", "B", base_font_size)
                    label = "Purpose — "
                    pdf.cell(pdf.get_string_width(sanitize_text(label)) + 1, line_h, sanitize_text(label), ln=0)
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size)
                    else:
                        pdf.set_font("Arial", "", base_font_size)
                    start_x = pdf.get_x()
                    used_width = start_x - (pdf.l_margin + INDENT)
                    pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(strip_md(body)))
                    pdf.set_xy(pdf.l_margin, pdf.get_y())
                    pdf.ln(0.5)
                    continue
                # STAT bullet — render as plain line (YoY: bold label prefix), then insert figure
                if content.upper().startswith("STAT") or content.upper().startswith("IMPACT"):
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size)
                    else:
                        pdf.set_font("Arial", "B", base_font_size)
                    left_margin = pdf.l_margin + INDENT
                    cur_y = pdf.get_y()
                    pdf.set_xy(left_margin, cur_y)
                    clean = re.sub(r"^(STAT|Impact)\s*(Impact)?\s*[:\-—]?\s*", "", content, flags=re.I)
                    clean = re.sub(r"Y1\s*[–-]?\s*Y6\s*:\s*", "Y1, Y2, Y3, Y4, Y5, Y6: ", clean)
                    clean = re.sub(r"Y1Y6\s*:\s*", "Y1, Y2, Y3, Y4, Y5, Y6: ", clean)
                    # YoY partial bold handling
                    did_split = False
                    if current_section and 'yoy' in current_section.lower():
                        m = re.match(r"^(\s*Y\s*1\s*,\s*Y\s*2\s*,\s*Y\s*3\s*,\s*Y\s*4\s*,\s*Y\s*5\s*,\s*Y\s*6\s*:)\s*(.*)$", strip_md(clean), flags=re.I)
                        if m:
                            label_txt = m.group(1).strip() + ' '
                            body_txt = m.group(2).strip()
                            try:
                                if UNICODE_FONT and UNICODE_FONT_BOLD:
                                    pdf.set_font("DejaVu", "B", base_font_size)
                                elif UNICODE_FONT:
                                    pdf.set_font("DejaVu", "", base_font_size + 1)
                                else:
                                    pdf.set_font("Arial", "B", base_font_size)
                            except Exception:
                                pdf.set_font("Arial", "B", base_font_size)
                            pdf.cell(pdf.get_string_width(sanitize_text(label_txt)) + 1, line_h, sanitize_text(label_txt), ln=0)
                            if UNICODE_FONT:
                                pdf.set_font("DejaVu", "", base_font_size)
                            else:
                                pdf.set_font("Arial", "", base_font_size)
                            start_x = pdf.get_x()
                            used_width = start_x - (pdf.l_margin + INDENT)
                            pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(body_txt))
                            did_split = True
                    if not did_split:
                        try:
                            if UNICODE_FONT and UNICODE_FONT_BOLD:
                                pdf.set_font("DejaVu", "B", base_font_size)
                            elif UNICODE_FONT:
                                pdf.set_font("DejaVu", "", base_font_size + 1)
                            else:
                                pdf.set_font("Arial", "B", base_font_size)
                        except Exception:
                            pdf.set_font("Arial", "B", base_font_size)
                        pdf.multi_cell(effective_page_width - INDENT, line_h, sanitize_text(strip_md(clean)))
                    pdf.set_xy(pdf.l_margin, pdf.get_y())
                    pdf.ln(0.5)
                    if not inserted_image_for_current:
                        inserted_image_for_current = _insert_section_image_for(current_section)
                    continue
                # Price Impact bullet — bullet + bold label
                if re.match(r'^(price\s*impact)\b', content, flags=re.I):
                    left_margin = pdf.l_margin + INDENT
                    cur_y = pdf.get_y()
                    pdf.set_xy(left_margin, cur_y)
                    pdf.cell(BULLET_CELL_W, line_h, BULLET, ln=0)
                    label = "Price Impact — "
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size)
                    else:
                        pdf.set_font("Arial", "B", base_font_size)
                    safe_label = sanitize_text(label)
                    pdf.cell(pdf.get_string_width(safe_label) + 1, line_h, safe_label, ln=0)
                    body = re.sub(r'^(price\s*impact)\s*[—:-]?\s*', '', content, flags=re.I)
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size)
                    else:
                        pdf.set_font("Arial", "", base_font_size)
                    start_x = pdf.get_x()
                    used_width = start_x - (pdf.l_margin + INDENT)
                    pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(strip_md(body)))
                    pdf.set_xy(pdf.l_margin, pdf.get_y())
                    pdf.ln(0.5)
                    continue

                # Generic bullets (Suggestions or other) — keep bullets aligned
                rest = strip_md(content)
                left_margin = pdf.l_margin + INDENT
                cur_y = pdf.get_y()
                pdf.set_xy(left_margin, cur_y)
                pdf.cell(BULLET_CELL_W, line_h, BULLET, ln=0)
                start_x = pdf.get_x()
                used_width = start_x - (pdf.l_margin + INDENT)
                pdf.multi_cell(effective_page_width - used_width, line_h, sanitize_text(rest))
                pdf.set_xy(pdf.l_margin, pdf.get_y())
                pdf.ln(0.5)
                continue

            # Fallback: handle "Title — subtitle" if not caught earlier
            if "—" in line and any(line.lower().startswith(n.lower()) for n in SECTION_TITLES):
                parts = [p.strip() for p in line.split("—", 1)]
                if len(parts) == 2:
                    title = title_with_emoji(strip_md(parts[0]))
                    subtitle = strip_md(parts[1])
                    pdf.ln(2)
                    pdf.set_text_color(80, 80, 80)
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size + 4)
                    else:
                        pdf.set_font("Arial", "B", base_font_size + 3)
                    pdf.multi_cell(effective_page_width, line_h + 2, sanitize_text(title))
                    if subtitle:
                        if UNICODE_FONT:
                            pdf.set_font("DejaVu", "", base_font_size)
                        else:
                            pdf.set_font("Arial", "I", base_font_size)
                        pdf.multi_cell(effective_page_width, line_h, sanitize_text(subtitle))
                    pdf.set_text_color(0, 0, 0)
                    if UNICODE_FONT:
                        pdf.set_font("DejaVu", "", base_font_size)
                    else:
                        pdf.set_font("Arial", "", base_font_size)
                    pdf.ln(1)
                    continue

            # Price Impact emphasis handled above

            # Generic non-bullet fallback paragraph
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(effective_page_width, line_h, sanitize_text(strip_md(line)))
            pdf.ln(1)

        # After finishing all lines, add a small separation
        pdf.ln(2)

    def save_fig_temp(fig, name):
        path = os.path.join(tempfile.gettempdir(), name)
        fig.savefig(path, bbox_inches="tight")
        return path

    img1 = save_fig_temp(plt.gcf(), "curr_plot.png")  # not used directly but keeps matplotlib happy
    img_infl = save_fig_temp(fig1, "inflation.png")
    img_shock = save_fig_temp(fig2, "shock.png")
    img_sim = save_fig_temp(fig3, "sim.png")

    logo_path = _find_logo_path()
    pdf = PDF(logo_path=logo_path)
    # Slightly larger margins to avoid right-edge overflow
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    # If unicode font exists, switch default body font to it (title font is handled in renderer)
    if UNICODE_FONT:
        try:
            pdf.add_font("DejaVu", "", UNICODE_FONT, uni=True)
            if UNICODE_FONT_BOLD:
                pdf.add_font("DejaVu", "B", UNICODE_FONT_BOLD, uni=True)
            pdf.set_font("DejaVu", "", 11)
        except Exception:
            pdf.set_font("Arial", size=11)
    else:
        pdf.set_font("Arial", size=11)

    effective_page_width = pdf.w - 2 * pdf.l_margin

    # Render the AI summary with preserved structure and embed matching figures inline
    section_images = {
        'yoy inflation': img_infl,
        'supply shock': img_shock,
        'shock stopper': img_shock,
        'monte carlo survivability': img_sim,
        'monte carlo': img_sim,
    }
    render_structured_summary(pdf, summary, effective_page_width, base_font_size=11, section_images=section_images)

    pdf.ln(4)
    pdf.set_font("Arial", "B", 12)
    pdf.multi_cell(0, 10, "Token Design by TDeFi")
    pdf.set_font("Arial", "", 11)
    tdefi_note = """
Token engineering is not about distribution schedules or supply caps alone — it aligns incentives, behavior, and long-term value creation. A well-engineered token system creates harmony between product usage, network growth, and token demand. Poor token design often leads to value leakage and unsustainable emissions. At TDeFi, we build token models as economic engines — driven by utility, governed by logic, and sustained by real adoption.

Designing for Demand, Not Just Distribution
A common pitfall is decoupling emissions from real demand. When supply outpaces utility, it triggers sell pressure and a negative flywheel. Our approach ties token releases to verifiable demand — active users, protocol revenue, and ecosystem participation — rewarding value creation, not just speculation.
""".strip()
    for para in tdefi_note.split("\n\n"):
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(effective_page_width, 8, sanitize_text(para))
        pdf.ln(1)

    out = pdf.output(dest="S")
    if isinstance(out, (bytes, bytearray)):
        pdf_bytes = bytes(out)
    else:
        pdf_bytes = out.encode("latin-1", "ignore")

    st.markdown("""
    <div style="text-align: center; margin: 1.2rem 0;">
        <h3>📄 Download Complete Report</h3>
    </div>
    """, unsafe_allow_html=True)
    st.download_button(
        "📄 Download PDF Report",
        data=pdf_bytes,
        file_name=f"{project_name}_Audit_Report.pdf",
        mime="application/pdf"
    )

# -----------------------------
# Footer
# -----------------------------
st.markdown("""
<div class="tdefi-powered">
    🚀 Powered by <strong>TDeFi</strong> — TradeDog Token Growth Studio
</div>
""", unsafe_allow_html=True)
